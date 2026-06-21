"""
History/diff rendering for the restock_diff (price / stock availability) processor.

A text diff is meaningless for restock watches - each history snapshot is just a short
string like "In Stock: True - Price: 12.34". So instead of a diff, this renders the whole
timeline as a simple smoothed line graph of price over time, coloured green where the item
was in stock and red where it was out of stock.

Conforms to processors.difference_base.DifferenceRenderer (module-level render()), and also
provides an optional get_data() hook served as JSON by /diff/<uuid>/processor-data so the
(potentially long) timeline stays out of the rendered HTML - same rationale as the preview
asset endpoint.
"""
import re
import statistics
import time

from flask_babel import gettext
from loguru import logger

# Snapshot format is written by processor.py:  f"In Stock: {in_stock} - Price: {price}"
_RE_PRICE = re.compile(r"Price:\s*([\d.]+)", re.IGNORECASE)
_RE_INSTOCK = re.compile(r"In Stock:\s*(True|False)", re.IGNORECASE)


def _parse_restock_snapshot(text):
    """Parse a snapshot string into (price: float|None, in_stock: bool|None)."""
    price = None
    in_stock = None
    if text:
        m = _RE_PRICE.search(text)
        if m:
            try:
                price = float(m.group(1))
            except (TypeError, ValueError):
                price = None
        mi = _RE_INSTOCK.search(text)
        if mi:
            in_stock = mi.group(1).lower() == 'true'
    return price, in_stock


def _currency(watch):
    try:
        return (watch.get('restock') or {}).get('currency') or ''
    except Exception:
        return ''


def _build_series(watch):
    """Read the full history into a [{timestamp, price, in_stock}] timeline (oldest -> newest)."""
    series = []
    for ts in list(watch.history.keys()):
        try:
            snapshot = watch.get_history_snapshot(timestamp=ts)
        except Exception as e:
            logger.error(f"Restock diff: unable to read snapshot {ts} for {watch.get('uuid')}: {e}")
            continue
        price, in_stock = _parse_restock_snapshot(snapshot)
        series.append({'timestamp': int(ts), 'price': price, 'in_stock': in_stock})
    return series


def compute_price_summary(series):
    """Compact price stats for the graph overlays + (future) watchlist deal-score badge.

    This is the single source of truth for "is the current price low/typical/high". It works
    off an already-built series (cheap, no extra I/O) so the same function can be reused at
    check time to cache the result on the watch and avoid scanning history at render time.

    Returns None when there are no prices.
    """
    prices = [p['price'] for p in series if p.get('price') is not None]
    if not prices:
        return None

    count = len(prices)
    current = prices[-1]                      # series is oldest -> newest, so last priced value
    ordered = sorted(prices)
    mn, mx = ordered[0], ordered[-1]
    avg = sum(prices) / count

    if count >= 2:
        try:
            q = statistics.quantiles(prices, n=4, method='inclusive')  # [p25, p50, p75]
            p25, median, p75 = q[0], q[1], q[2]
        except statistics.StatisticsError:
            median, p25, p75 = statistics.median(prices), mn, mx
    else:
        median = p25 = p75 = current

    # Two directional shares so the UI can phrase it naturally per status:
    #   cheaper_than_pct = share of records MORE expensive than now (good when price is low)
    #   pricier_than_pct = share of records CHEAPER than now (relevant when price is high)
    cheaper_than_pct = round(100.0 * sum(1 for x in prices if x > current) / count)
    pricier_than_pct = round(100.0 * sum(1 for x in prices if x < current) / count)

    if current <= p25:
        status = 'low'
    elif current >= p75:
        status = 'high'
    else:
        status = 'typical'

    r2 = lambda v: round(v, 2)
    return {
        'count': count,
        'min': r2(mn), 'max': r2(mx), 'avg': r2(avg), 'median': r2(median),
        'p25': r2(p25), 'p75': r2(p75),
        'current': r2(current),
        'cheaper_than_pct': cheaper_than_pct,
        'pricier_than_pct': pricier_than_pct,
        'status': status,
        'all_time_low': current <= mn,
    }


def export_xlsx(watch, datastore):
    """Build an .xlsx of the full price/stock history. Returns (bytes, filename).

    xlsx (not csv) so prices stay real numbers and dates stay real dates in the user's
    spreadsheet, regardless of locale number formatting. Columns: Date, Stock status, Price,
    Currency. Served by /diff/<uuid>/processor-export.xlsx.
    """
    import datetime
    from io import BytesIO
    from openpyxl import Workbook

    series = _build_series(watch)
    currency = _currency(watch)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Price history'
    ws.append([gettext('Date'), gettext('Stock status'), gettext('Price'), gettext('Currency')])

    for p in series:
        when = datetime.datetime.fromtimestamp(p['timestamp'])
        if p['in_stock'] is None:
            stock = ''
        else:
            stock = gettext('In stock') if p['in_stock'] else gettext('Out of stock')
        cell_date = ws.cell(row=ws.max_row + 1, column=1, value=when)
        cell_date.number_format = 'YYYY-MM-DD HH:MM:SS'
        ws.cell(row=ws.max_row, column=2, value=stock)
        # Real numeric price (None -> blank cell) so spreadsheet number formats apply.
        ws.cell(row=ws.max_row, column=3, value=p['price'])
        ws.cell(row=ws.max_row, column=4, value=currency)

    # Sensible column widths.
    for col, width in {'A': 20, 'B': 14, 'C': 12, 'D': 10}.items():
        ws.column_dimensions[col].width = width

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), f"price-history-{watch.get('uuid')}.xlsx"


def get_data(watch, datastore, request):
    """JSON payload for the price/stock graph, fetched via /diff/<uuid>/processor-data.
    Keeps the full timeline out of the HTML page."""
    series = _build_series(watch)
    priced = sum(1 for p in series if p['price'] is not None)
    logger.info(f"Restock diff get_data for {watch.get('uuid')}: {len(series)} snapshots, {priced} with a price")
    return {
        'series': series,
        'currency': _currency(watch),
        'summary': compute_price_summary(series),
    }


def render(watch, datastore, request, url_for, render_template, flash, redirect, extract_form=None):
    """Render the price/stock timeline page (shell + summary). The graph data is loaded
    asynchronously from get_data() so this stays light regardless of history length."""
    uuid = watch.get('uuid')
    dates = list(watch.history.keys())

    # Light render: read only the most recent snapshot for the summary badge/price.
    latest = None
    if dates:
        try:
            price, in_stock = _parse_restock_snapshot(watch.get_history_snapshot(timestamp=dates[-1]))
            latest = {'timestamp': int(dates[-1]), 'price': price, 'in_stock': in_stock}
        except Exception as e:
            logger.error(f"Restock diff: unable to read latest snapshot for {uuid}: {e}")

    # Opening the history page counts as viewing it (mirrors the text diff page).
    datastore.set_last_viewed(uuid, time.time())

    return render_template(
        'restock_diff/difference.html',
        uuid=uuid,
        watch=watch,
        current_diff_url=watch['url'],
        extra_title=f" - {watch.label} - {gettext('Price history')}",
        last_error=watch['last_error'],
        last_error_screenshot=watch.get_error_snapshot(),
        last_error_text=watch.get_error_text(),
        versions=dates,
        from_version=str(dates[-2]) if len(dates) >= 2 else (str(dates[-1]) if dates else ''),
        to_version=str(dates[-1]) if dates else '',
        restock_latest=latest,
        restock_currency=_currency(watch),
        has_enough_history=len(dates) >= 2,
        processor_data_url=url_for('ui.ui_diff.diff_history_page_processor_data', uuid=uuid),
    )
