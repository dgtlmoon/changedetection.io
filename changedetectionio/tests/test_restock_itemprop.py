#!/usr/bin/env python3
import os
import time

from flask import url_for
from .util import live_server_setup, wait_for_all_checks, wait_for_notification_endpoint_output, extract_UUID_from_client, delete_all_watches
from ..notification import default_notification_format

instock_props = [
    # LD+JSON with non-standard list of 'type' https://github.com/dgtlmoon/changedetection.io/issues/1833
    '<script type=\'application/ld+json\'>{"@context": "http://schema.org","@type": ["Product", "SubType"],"name": "My test product","description":"","Offers": {    "@type": "Offer",    "offeredBy": {        "@type": "Organization",        "name":"Person",       "telephone":"+1 999 999 999"    },    "price": $$PRICE$$,    "priceCurrency": "EUR",    "url": "/some/url", "availability": "http://schema.org/InStock"}        }</script>',
    # LD JSON
    '<script id="product-jsonld" type="application/ld+json">{"@context":"https://schema.org","@type":"Product","brand":{"@type":"Brand","name":"Ubiquiti"},"name":"UniFi Express","sku":"UX","description":"Impressively compact UniFi Cloud Gateway and WiFi 6 access point that runs UniFi Network. Powers an entire network or simply meshes as an access point.","url":"https://store.ui.com/us/en/products/ux","image":{"@type":"ImageObject","url":"https://cdn.ecomm.ui.com/products/4ed25b4c-db92-4b98-bbf3-b0989f007c0e/123417a2-895e-49c7-ba04-b6cd8f6acc03.png","width":"1500","height":"1500"},"offers":{"@type":"Offer","availability":"https://schema.org/InStock","priceSpecification":{"@type":"PriceSpecification","price":$$PRICE$$,"priceCurrency":"USD","valueAddedTaxIncluded":false}}}</script>',
    '<script id="product-schema" type="application/ld+json">{"@context": "https://schema.org","@type": "Product","itemCondition": "https://schema.org/NewCondition","image": "//1.com/hmgo","name": "Polo MuscleFit","color": "Beige","description": "Polo","sku": "0957102010","brand": {"@type": "Brand","name": "H&M"},"category": {"@type": "Thing","name": "Polo"},"offers": [{"@type": "Offer","url": "https:/www2.xxxxxx.com/fr_fr/productpage.0957102010.html","priceCurrency": "EUR","price": $$PRICE$$,"availability": "http://schema.org/InStock","seller": {  "@type": "Organization", "name": "H&amp;M"}}]}</script>'
    # Microdata
    '<div itemscope itemtype="https://schema.org/Product"><h1 itemprop="name">Example Product</h1><p itemprop="description">This is a sample product description.</p><div itemprop="offers" itemscope itemtype="https://schema.org/Offer"><p>Price: <span itemprop="price">$$$PRICE$$</span></p><link itemprop="availability" href="https://schema.org/InStock" /></div></div>'
]

out_of_stock_props = [
    # out of stock AND contains multiples
    '<script type="application/ld+json">{"@context":"http://schema.org","@type":"WebSite","url":"https://www.medimops.de/","potentialAction":{"@type":"SearchAction","target":"https://www.medimops.de/produkte-C0/?fcIsSearch=1&searchparam={searchparam}","query-input":"required name=searchparam"}}</script><script type="application/ld+json">{"@context":"http://schema.org","@type":"Product","name":"Horsetrader: Robert Sangster and the Rise and Fall of the Sport of Kings","image":"https://images2.medimops.eu/product/43a982/M00002551322-large.jpg","productID":"isbn:9780002551328","gtin13":"9780002551328","category":"Livres en langue étrangère","offers":{"@type":"Offer","priceCurrency":"EUR","price":$$PRICE$$,"itemCondition":"UsedCondition","availability":"OutOfStock"},"brand":{"@type":"Thing","name":"Patrick Robinson","url":"https://www.momox-shop.fr/,patrick-robinson/"}}</script>'
]

def set_original_response(datastore_path, props_markup='', price="121.95"):
    props_markup=props_markup.replace('$$PRICE$$', price)
    test_return_data = f"""<html>
       <body>
     Some initial text<br>
     <p>Which is across multiple lines</p>
     <br>
     So let's see what happens.  <br>
     <div>price: ${price}</div>
     {props_markup}
     </body>
     </html>
    """

    with open(os.path.join(datastore_path, "endpoint-content.txt"), "w") as f:
        f.write(test_return_data)
    time.sleep(1)
    return None

def test_restock_itemprop_basic(client, live_server, measure_memory_usage, datastore_path):

    test_url = url_for('test_endpoint', _external=True)

    # By default it should enable ('in_stock_processing') == 'all_changes'

    for p in instock_props:
        set_original_response(props_markup=p, datastore_path=datastore_path)
        client.post(
            url_for("ui.ui_views.form_quick_watch_add"),
            data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
            follow_redirects=True
        )
        wait_for_all_checks(client)
        res = client.get(url_for("watchlist.index"))
        assert b'more than one price detected' not in res.data
        assert b'has-restock-info' in res.data
        assert b' in-stock' in res.data
        assert b' not-in-stock' not in res.data
        delete_all_watches(client)


    for p in out_of_stock_props:
        set_original_response(props_markup=p, datastore_path=datastore_path)
        client.post(
            url_for("ui.ui_views.form_quick_watch_add"),
            data={"url": test_url, "tags": '', 'processor': 'restock_diff'},
            follow_redirects=True
        )
        wait_for_all_checks(client)
        res = client.get(url_for("watchlist.index"))

        assert b'has-restock-info not-in-stock' in res.data

        delete_all_watches(client)

def test_itemprop_price_change(client, live_server, measure_memory_usage, datastore_path):
    

    # Out of the box 'Follow price changes' should be ON
    test_url = url_for('test_endpoint', _external=True)

    set_original_response(props_markup=instock_props[0], price="190.95", datastore_path=datastore_path)
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )

    # A change in price, should trigger a change by default
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'190.95' in res.data

    # basic price change, look for notification
    set_original_response(props_markup=instock_props[0], price='180.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'180.45' in res.data
    assert b'has-unread-changes' in res.data
    client.get(url_for("ui.mark_all_viewed"), follow_redirects=True)
    time.sleep(0.2)


    # turning off price change trigger, but it should show the new price, with no change notification
    set_original_response(props_markup=instock_props[0], price='120.45', datastore_path=datastore_path)
    res = client.post(
        url_for("ui.ui_edit.edit_page", uuid="first"),
        data={"processor_config_restock_diff-follow_price_changes": "", "url": test_url, "tags": "", "headers": "", 'fetch_backend': "html_requests", "time_between_check_use_default": "y"},
        follow_redirects=True
    )
    assert b"Updated watch." in res.data
    client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'120.45' in res.data
    assert b'has-unread-changes' not in res.data


    delete_all_watches(client)

def _run_test_minmax_limit(client, extra_watch_edit_form, datastore_path):

    delete_all_watches(client)

    test_url = url_for('test_endpoint', _external=True)

    set_original_response(props_markup=instock_props[0], price="950.95", datastore_path=datastore_path)
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )
    wait_for_all_checks(client)

    data = {
        "tags": "",
        "url": test_url,
        "headers": "",
        "time_between_check-hours": 5,
        'fetch_backend': "html_requests",
        "time_between_check_use_default": "y"
    }
    data.update(extra_watch_edit_form)
    res = client.post(
        url_for("ui.ui_edit.edit_page", uuid="first"),
        data=data,
        follow_redirects=True
    )
    assert b"Updated watch." in res.data
    wait_for_all_checks(client)

    client.get(url_for("ui.mark_all_viewed"))

    # price changed to something greater than min (900), BUT less than max (1100).. should be no change
    set_original_response(props_markup=instock_props[0], price='1000.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"))
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))

    assert b'more than one price detected' not in res.data
    # BUT the new price should show, even tho its within limits
    assert b'1,000.45' or b'1000.45' in res.data #depending on locale
    assert b'has-unread-changes' not in res.data

    # price changed to something LESS than min (900), SHOULD be a change
    set_original_response(props_markup=instock_props[0], price='890.45', datastore_path=datastore_path)

    res = client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)
    assert b'Queued 1 watch for rechecking.' in res.data
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'890.45' in res.data
    assert b'has-unread-changes' in res.data

    client.get(url_for("ui.mark_all_viewed"))


    # 2715 - Price detection (once it crosses the "lower" threshold) again with a lower price - should trigger again!
    set_original_response(props_markup=instock_props[0], price='820.45', datastore_path=datastore_path)
    res = client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)
    assert b'Queued 1 watch for rechecking.' in res.data
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'820.45' in res.data
    assert b'has-unread-changes' in res.data
    client.get(url_for("ui.mark_all_viewed"))

    # price changed to something MORE than max (1100.10), SHOULD be a change
    set_original_response(props_markup=instock_props[0], price='1890.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    # Depending on the LOCALE it may be either of these (generally for US/default/etc)
    assert b'1,890.45' in res.data or b'1890.45' in res.data
    assert b'has-unread-changes' in res.data

    delete_all_watches(client)


def test_restock_itemprop_minmax(client, live_server, measure_memory_usage, datastore_path):
    
    extras = {
        "processor_config_restock_diff-follow_price_changes": "y",
        "processor_config_restock_diff-price_change_min": 900.0,
        "processor_config_restock_diff-price_change_max": 1100.10
    }
    _run_test_minmax_limit(client, extra_watch_edit_form=extras, datastore_path=datastore_path)

def test_restock_itemprop_with_tag(client, live_server, measure_memory_usage, datastore_path):
    

    res = client.post(
        url_for("tags.form_tag_add"),
        data={"name": "test-tag"},
        follow_redirects=True
    )
    assert b"Tag added" in res.data

    res = client.post(
        url_for("tags.form_tag_edit_submit", uuid="first"),
        data={"name": "test-tag",
              "processor_config_restock_diff-follow_price_changes": "y",
              "processor_config_restock_diff-price_change_min": 900.0,
              "processor_config_restock_diff-price_change_max": 1100.10,
              "overrides_watch": "y", #overrides_watch should be restock_overrides_watch
              },
        follow_redirects=True
    )

    extras = {
        "tags": "test-tag"
    }

    _run_test_minmax_limit(client, extra_watch_edit_form=extras,datastore_path=datastore_path)
    delete_all_watches(client)



def test_itemprop_percent_threshold(client, live_server, measure_memory_usage, datastore_path):

    delete_all_watches(client)

    test_url = url_for('test_endpoint', _external=True)

    set_original_response(props_markup=instock_props[0], price="950.95", datastore_path=datastore_path)
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )

    # A change in price, should trigger a change by default
    wait_for_all_checks(client)

    res = client.post(
        url_for("ui.ui_edit.edit_page", uuid="first"),
        data={"processor_config_restock_diff-follow_price_changes": "y",
              "processor_config_restock_diff-price_change_threshold_percent": 5.0,
              "url": test_url,
              "tags": "",
              "headers": "",
              'fetch_backend': "html_requests",
              "time_between_check_use_default": "y"
              },
        follow_redirects=True
    )
    assert b"Updated watch." in res.data
    wait_for_all_checks(client)


    # Basic change should not trigger
    set_original_response(props_markup=instock_props[0], price='960.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"))
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'960.45' in res.data
    assert b'has-unread-changes' not in res.data

    # Bigger INCREASE change than the threshold should trigger
    set_original_response(props_markup=instock_props[0], price='1960.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"))
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'1,960.45' in res.data or b'1960.45' in res.data #depending on locale
    assert b'has-unread-changes' in res.data


    # Small decrease should NOT trigger
    client.get(url_for("ui.mark_all_viewed"))
    set_original_response(props_markup=instock_props[0], price='1950.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"))
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'1,950.45' in res.data or b'1950.45' in res.data #depending on locale
    assert b'has-unread-changes' not in res.data

    # PROOF that the threshold is measured "since the PREVIOUS check" and NOT "since the first check":
    # a slow upward creep where every single step is below the 5% threshold versus the *previous*
    # check, but the total drift from where the creep started (1950.45) ends up ABOVE 5%.
    #   1950.45 -> 2000.00  = +2.54% vs previous  (below 5%)
    #   2000.00 -> 2050.00  = +2.50% vs previous  (below 5%)
    #   2050.00 -> 2100.00  = +2.44% vs previous  (below 5%)
    #   1950.45 -> 2100.00  = +7.67% in total     (ABOVE 5%)
    # Under "since previous check" NONE of these trigger (each step is sub-threshold).
    # Under "since first check" the accumulated drift would cross 5% and trigger here - so the
    # final assertion below would fail. We deliberately never mark_all_viewed during the creep,
    # so any single trigger would leave has-unread-changes set.
    for creep_price in ['2000.00', '2050.00', '2100.00']:
        set_original_response(props_markup=instock_props[0], price=creep_price, datastore_path=datastore_path)
        client.get(url_for("ui.form_watch_checknow"))
        wait_for_all_checks(client)

    res = client.get(url_for("watchlist.index"))
    assert b'2,100.00' in res.data or b'2100.00' in res.data #depending on locale
    # +7.67% total drift since the creep started, yet still unread-free -> comparison is vs PREVIOUS check
    assert b'has-unread-changes' not in res.data


    # Re #2600 - Switch the mode to normal type and back, and see if the values stick..
    ###################################################################################
    uuid = next(iter(live_server.app.config['DATASTORE'].data['watching']))

    res = client.post(
        url_for("ui.ui_edit.edit_page", uuid=uuid),
        data={"processor_config_restock_diff-follow_price_changes": "y",
              "processor_config_restock_diff-price_change_threshold_percent": 5.05,
              "processor": "text_json_diff",
              "url": test_url,
              'fetch_backend': "html_requests",
              "time_between_check_use_default": "y"
              },
        follow_redirects=True
    )
    assert b"Updated watch." in res.data
    # And back again
    live_server.app.config['DATASTORE'].data['watching'][uuid]['processor'] = 'restock_diff'
    res = client.get(url_for("ui.ui_edit.edit_page", uuid=uuid))
    assert b'type="text" value="5.05"' in res.data

    delete_all_watches(client)



def test_change_with_notification_values(client, live_server, measure_memory_usage, datastore_path):
    

    if os.path.isfile(os.path.join(datastore_path, "notification.txt")):
        os.unlink(os.path.join(datastore_path, "notification.txt"))

    test_url = url_for('test_endpoint', _external=True)
    set_original_response(props_markup=instock_props[0], price='960.45', datastore_path=datastore_path)

    notification_url = url_for('test_notification_endpoint', _external=True).replace('http', 'json')

    ######################
    # You must add a type of 'restock_diff' for its tokens to register as valid in the global settings
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )

    # A change in price, should trigger a change by default
    wait_for_all_checks(client)

    # Should see new tokens register — the placeholder table lives on the
    # notifications page now (post-/settings refactor).
    res = client.get(url_for("settings.notifications.apprise"))

    assert b'{{restock.last_price}}' in res.data
    assert b'{{restock.previous_price}}' in res.data
    assert b'Price at the previous check' in res.data

    #####################
    # Set this up for when we remove the notification from the watch, it should fallback with these details
    res = client.post(
        url_for("settings.notifications.apprise"),
        data={"notification_urls": notification_url,
              "notification_title": "title new price {{restock.price}}",
              "notification_body": "new price {{restock.price}} previous price {{restock.previous_price}} instock {{restock.in_stock}}",
              "notification_format": default_notification_format},
        follow_redirects=True
    )

    # check tag accepts without error

    # Check the watches in these modes add the tokens for validating
    assert b"A variable or function is not defined" not in res.data

    assert b"Settings updated." in res.data

    # A change in price, should trigger a change by default
    set_original_response(props_markup=instock_props[0], price='1950.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"))
    wait_for_all_checks(client)
    wait_for_notification_endpoint_output(datastore_path=datastore_path)
    assert os.path.isfile(os.path.join(datastore_path, "notification.txt")), "Notification received"
    with open(os.path.join(datastore_path, "notification.txt"), 'r') as f:
        notification = f.read()
        assert "new price 1950.45" in notification
        assert "title new price 1950.45" in notification
        assert "previous price 960.45" in notification

    # Regression for #4260: with 3+ snapshots, {{restock.previous_price}} must report the price at
    # the PREVIOUS check (1950.45), not the first-ever price in history (960.45). Before the fix the
    # newest-first history list was indexed at [-1] (the oldest snapshot), so from the 3rd check on
    # this was stuck on the first price ever recorded. The two-snapshot case above passed only by
    # coincidence (with two entries the oldest snapshot IS the previous check).
    os.unlink(os.path.join(datastore_path, "notification.txt"))
    set_original_response(props_markup=instock_props[0], price='2500.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"))
    wait_for_all_checks(client)
    wait_for_notification_endpoint_output(datastore_path=datastore_path)
    assert os.path.isfile(os.path.join(datastore_path, "notification.txt")), "Notification received"
    with open(os.path.join(datastore_path, "notification.txt"), 'r') as f:
        notification = f.read()
        assert "new price 2500.45" in notification
        assert "previous price 1950.45" in notification          # the actual previous check
        assert "previous price 960.45" not in notification        # the pre-fix bug (first-ever price)

    ## Now test the "SEND TEST NOTIFICATION" is working
    os.unlink(os.path.join(datastore_path, "notification.txt"))
    uuid = next(iter(live_server.app.config['DATASTORE'].data['watching']))
    res = client.post(url_for("ui.ui_notification.ajax_callback_send_notification_test", watch_uuid=uuid), data={}, follow_redirects=True)
    wait_for_notification_endpoint_output(datastore_path=datastore_path)
    assert os.path.isfile(os.path.join(datastore_path, "notification.txt")), "Notification received"

    delete_all_watches(client)

def test_data_sanity(client, live_server, measure_memory_usage, datastore_path):
    

    delete_all_watches(client)

    test_url = url_for('test_endpoint', _external=True)
    test_url2 = url_for('test_endpoint2', _external=True)
    set_original_response(props_markup=instock_props[0], price="950.95", datastore_path=datastore_path)
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )

    client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)

    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert b'950.95' in res.data

    # Check the restock model object doesnt store the value by mistake and used in a new one
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url2, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )
    client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)
    wait_for_all_checks(client)
    res = client.get(url_for("watchlist.index"))
    assert str(res.data.decode()).count("950.95") == 1, "Price should only show once (for the watch added, no other watches yet)"

    ## different test, check the edit page works on an empty request result
    delete_all_watches(client)

    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url2, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )
    wait_for_all_checks(client)

    res = client.get(
        url_for("ui.ui_edit.edit_page", uuid="first"))
    assert test_url2.encode('utf-8') in res.data

    delete_all_watches(client)

# All examples should give a prive of 666.66
def test_special_prop_examples(client, live_server, measure_memory_usage, datastore_path):
    import glob
    

    test_url = url_for('test_endpoint', _external=True)
    check_path = os.path.join(os.path.dirname(__file__), "itemprop_test_examples", "*.txt")
    files = glob.glob(check_path)
    assert files
    for test_example_filename in files:
        with open(test_example_filename, 'r') as example_f:
            with open(os.path.join(datastore_path, "endpoint-content.txt"), "w") as test_f:
                test_f.write(f"<html><body>{example_f.read()}</body></html>")

            # Now fetch it and check the price worked
            client.post(
                url_for("ui.ui_views.form_quick_watch_add"),
                data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
                follow_redirects=True
            )
            wait_for_all_checks(client)
            res = client.get(url_for("watchlist.index"))
            assert b'ception' not in res.data
            assert b'155.55' in res.data

    delete_all_watches(client)


def test_itemprop_as_str(client, live_server, measure_memory_usage, datastore_path):

    test_return_data = f"""<html>
       <body>
     Some initial text<br>
     <p>Which is across multiple lines</p>
<span itemprop="offers" itemscope itemtype="http://schema.org/Offer">
<meta content="767.55" itemprop="price"/>
<meta content="EUR" itemprop="priceCurrency"/>
<meta content="InStock" itemprop="availability"/>
<meta content="https://www.123-test.dk" itemprop="url"/>
</span>
     </body>
     </html>
    """

    with open(os.path.join(datastore_path, "endpoint-content.txt"), "w") as f:
        f.write(test_return_data)


    test_url = url_for('test_endpoint', _external=True)

    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": 'restock tests', 'processor': 'restock_diff'},
        follow_redirects=True
    )

    client.get(url_for("ui.form_watch_checknow"))
    wait_for_all_checks(client)

    res = client.get(url_for("watchlist.index"))
    assert b'767.55' in res.data


def test_restock_diff_price_data_ajax(client, live_server, measure_memory_usage, datastore_path):
    """The restock history graph fetches its timeline from the processor-data callback
    (restock_diff/difference.py::get_data, served at /diff/<uuid>/processor-data)."""
    import json

    test_url = url_for('test_endpoint', _external=True)

    # First snapshot - in stock @ 190.95
    set_original_response(props_markup=instock_props[0], price="190.95", datastore_path=datastore_path)
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": '', 'processor': 'restock_diff'},
        follow_redirects=True
    )
    wait_for_all_checks(client)
    uuid = extract_UUID_from_client(client)

    # Second snapshot - price changes to 180.45 (still in stock) -> a new history point
    set_original_response(props_markup=instock_props[0], price='180.45', datastore_path=datastore_path)
    client.get(url_for("ui.form_watch_checknow"), follow_redirects=True)
    wait_for_all_checks(client)

    # The AJAX/plugin data callback
    res = client.get(url_for("ui.ui_diff.diff_history_page_processor_data", uuid=uuid))
    assert res.status_code == 200
    data = json.loads(res.data)

    assert 'series' in data
    assert 'currency' in data
    series = data['series']
    assert len(series) >= 2

    # Every point exposes the parsed timestamp + price + stock state
    for point in series:
        assert 'timestamp' in point
        assert 'price' in point
        assert 'in_stock' in point

    prices = [p['price'] for p in series]
    assert 190.95 in prices
    assert 180.45 in prices
    # These snapshots were all "in stock"
    assert series[-1]['in_stock'] is True

    # Price summary (deal-score inputs) is computed and returned
    assert data.get('summary')
    summary = data['summary']
    for key in ('count', 'min', 'max', 'avg', 'median', 'p25', 'p75', 'current', 'status', 'all_time_low'):
        assert key in summary
    assert summary['status'] in ('low', 'typical', 'high')
    assert summary['min'] <= summary['avg'] <= summary['max']

    # XLSX export of the history
    res = client.get(url_for("ui.ui_diff.diff_history_page_processor_export", uuid=uuid))
    assert res.status_code == 200
    assert res.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert 'attachment' in res.headers.get('Content-Disposition', '')
    # Valid xlsx with the expected header row + numeric prices
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(res.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ('Date', 'Stock status', 'Price', 'Currency')
    xlsx_prices = [r[2] for r in rows[1:]]
    assert 190.95 in xlsx_prices
    assert 180.45 in xlsx_prices

    delete_all_watches(client)


def test_restock_edit_has_ai_llm_section(client, live_server, measure_memory_usage, datastore_path):
    """restock_diff watches must expose the AI / LLM tab content on the edit page, the same as
    text_json_diff. Regression: the AI section used to be gated to text_json_diff only, so the
    #ai-llm tab rendered empty for restock watches."""
    test_url = url_for('test_endpoint', _external=True)
    set_original_response(props_markup=instock_props[0], datastore_path=datastore_path)
    client.post(
        url_for("ui.ui_views.form_quick_watch_add"),
        data={"url": test_url, "tags": '', 'processor': 'restock_diff'},
        follow_redirects=True
    )
    wait_for_all_checks(client)
    uuid = extract_UUID_from_client(client)

    res = client.get(url_for("ui.ui_edit.edit_page", uuid=uuid))
    assert res.status_code == 200
    # The AI / LLM tab link exists...
    assert b'#ai-llm' in res.data
    # ...and crucially its section now renders for restock_diff (either the configured fields
    # block or the "configure a provider" disabled block — both carry this id). Before the fix
    # this was absent for restock watches.
    assert b'llm-intent-section' in res.data

    delete_all_watches(client)