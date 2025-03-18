from abc import abstractmethod
import time
from wtforms import ValidationError
from loguru import logger

from changedetectionio.forms import validate_url


class Importer():
    remaining_data = []
    new_uuids = []
    good = 0

    def __init__(self):
        self.new_uuids = []
        self.good = 0
        self.remaining_data = []
        self.import_profile = None

    @abstractmethod
    def run(self,
            data,
            flash,
            datastore):
        pass


class import_url_list(Importer):
    """
    Imports a list, can be in <code>https://example.com tag1, tag2, last tag</code> format
    """
    def run(self,
            data,
            flash,
            datastore,
            processor=None
            ):

        urls = data.split("\n")
        good = 0
        now = time.time()

        if (len(urls) > 5000):
            flash("Importing 5,000 of the first URLs from your list, the rest can be imported again.")

        for url in urls:
            url = url.strip()
            if not len(url):
                continue

            tags = ""

            # 'tags' should be a csv list after the URL
            if ' ' in url:
                url, tags = url.split(" ", 1)

            # Flask wtform validators wont work with basic auth, use validators package
            # Up to 5000 per batch so we dont flood the server
            # @todo validators.url will fail when you add your own IP etc
            if len(url) and 'http' in url.lower() and good < 5000:
                extras = None
                if processor:
                    extras = {'processor': processor}
                new_uuid = datastore.add_watch(url=url.strip(), tag=tags, write_to_disk_now=False, extras=extras)

                if new_uuid:
                    # Straight into the queue.
                    self.new_uuids.append(new_uuid)
                    good += 1
                    continue

            # Worked past the 'continue' above, append it to the bad list
            if self.remaining_data is None:
                self.remaining_data = []
            self.remaining_data.append(url)

        flash("{} Imported from list in {:.2f}s, {} Skipped.".format(good, time.time() - now, len(self.remaining_data)))


class import_distill_io_json(Importer):
    def run(self,
            data,
            flash,
            datastore,
            ):

        import json
        good = 0
        now = time.time()
        self.new_uuids=[]

        # @todo Use JSONSchema like in the API to validate here.
        
        try:
            data = json.loads(data.strip())
        except json.decoder.JSONDecodeError:
            flash("Unable to read JSON file, was it broken?", 'error')
            return

        if not data.get('data'):
            flash("JSON structure looks invalid, was it broken?", 'error')
            return

        for d in data.get('data'):
            d_config = json.loads(d['config'])
            extras = {'title': d.get('name', None)}

            if len(d['uri']) and good < 5000:
                try:
                    # @todo we only support CSS ones at the moment
                    if d_config['selections'][0]['frames'][0]['excludes'][0]['type'] == 'css':
                        extras['subtractive_selectors'] = d_config['selections'][0]['frames'][0]['excludes'][0]['expr']
                except KeyError:
                    pass
                except IndexError:
                    pass
                extras['include_filters'] = []
                try:
                    if d_config['selections'][0]['frames'][0]['includes'][0]['type'] == 'xpath':
                        extras['include_filters'].append('xpath:' + d_config['selections'][0]['frames'][0]['includes'][0]['expr'])
                    else:
                        extras['include_filters'].append(d_config['selections'][0]['frames'][0]['includes'][0]['expr'])
                except KeyError:
                    pass
                except IndexError:
                    pass

                new_uuid = datastore.add_watch(url=d['uri'].strip(),
                                               tag=",".join(d.get('tags', [])),
                                               extras=extras,
                                               write_to_disk_now=False)

                if new_uuid:
                    # Straight into the queue.
                    self.new_uuids.append(new_uuid)
                    good += 1

        flash("{} Imported from Distill.io in {:.2f}s, {} Skipped.".format(len(self.new_uuids), time.time() - now, len(self.remaining_data)))


class import_xlsx_wachete(Importer):

    def run(self,
            data,
            flash,
            datastore,
            ):

        good = 0
        now = time.time()
        self.new_uuids = []

        from openpyxl import load_workbook

        try:
            wb = load_workbook(data)
        except Exception as e:
            # @todo correct except
            flash("Unable to read export XLSX file, something wrong with the file?", 'error')
            return

        row_id = 2
        for row in wb.active.iter_rows(min_row=row_id):
            try:
                extras = {}
                data = {}
                for cell in row:
                    if not cell.value:
                        continue
                    column_title = wb.active.cell(row=1, column=cell.column).value.strip().lower()
                    data[column_title] = cell.value

                # Forced switch to webdriver/playwright/etc
                dynamic_wachet = str(data.get('dynamic wachet', '')).strip().lower()  # Convert bool to str to cover all cases
                # libreoffice and others can have it as =FALSE() =TRUE(), or bool(true)
                if 'true' in dynamic_wachet or dynamic_wachet == '1':
                    extras['fetch_backend'] = 'html_webdriver'
                elif 'false' in dynamic_wachet or dynamic_wachet == '0':
                    extras['fetch_backend'] = 'html_requests'

                if data.get('xpath'):
                    # @todo split by || ?
                    extras['include_filters'] = [data.get('xpath')]
                if data.get('name'):
                    extras['title'] = data.get('name').strip()
                if data.get('interval (min)'):
                    minutes = int(data.get('interval (min)'))
                    hours, minutes = divmod(minutes, 60)
                    days, hours = divmod(hours, 24)
                    weeks, days = divmod(days, 7)
                    extras['time_between_check'] = {'weeks': weeks, 'days': days, 'hours': hours, 'minutes': minutes, 'seconds': 0}

                # At minimum a URL is required.
                if data.get('url'):
                    try:
                        validate_url(data.get('url'))
                    except ValidationError as e:
                        logger.error(f">> Import URL error {data.get('url')} {str(e)}")
                        flash(f"Error processing row number {row_id}, URL value was incorrect, row was skipped.", 'error')
                        # Don't bother processing anything else on this row
                        continue

                    new_uuid = datastore.add_watch(url=data['url'].strip(),
                                                   extras=extras,
                                                   tag=data.get('folder'),
                                                   write_to_disk_now=False)
                    if new_uuid:
                        # Straight into the queue.
                        self.new_uuids.append(new_uuid)
                        good += 1
            except Exception as e:
                logger.error(e)
                flash(f"Error processing row number {row_id}, check all cell data types are correct, row was skipped.", 'error')
            else:
                row_id += 1

        flash(
            "{} imported from Wachete .xlsx in {:.2f}s".format(len(self.new_uuids), time.time() - now))


class import_xlsx_custom(Importer):

    def run(self,
            data,
            flash,
            datastore,
            ):

        good = 0
        now = time.time()
        self.new_uuids = []

        from openpyxl import load_workbook

        try:
            wb = load_workbook(data)
        except Exception as e:
            # @todo correct except
            flash("Unable to read export XLSX file, something wrong with the file?", 'error')
            return

        # @todo cehck atleast 2 rows, same in other method
        from changedetectionio.forms import validate_url
        row_i = 1

        try:
            for row in wb.active.iter_rows():
                url = None
                tags = None
                extras = {}

                for cell in row:
                    if not self.import_profile.get(cell.col_idx):
                        continue
                    if not cell.value:
                        continue

                    cell_map = self.import_profile.get(cell.col_idx)

                    cell_val = str(cell.value).strip()  # could be bool

                    if cell_map == 'url':
                        url = cell.value.strip()
                        try:
                            validate_url(url)
                        except ValidationError as e:
                            logger.error(f">> Import URL error {url} {str(e)}")
                            flash(f"Error processing row number {row_i}, URL value was incorrect, row was skipped.", 'error')
                            # Don't bother processing anything else on this row
                            url = None
                            break
                    elif cell_map == 'tag':
                        tags = cell.value.strip()
                    elif cell_map == 'include_filters':
                        # @todo validate?
                        extras['include_filters'] = [cell.value.strip()]
                    elif cell_map == 'interval_minutes':
                        hours, minutes = divmod(int(cell_val), 60)
                        days, hours = divmod(hours, 24)
                        weeks, days = divmod(days, 7)
                        extras['time_between_check'] = {'weeks': weeks, 'days': days, 'hours': hours, 'minutes': minutes, 'seconds': 0}
                    else:
                        extras[cell_map] = cell_val

                # At minimum a URL is required.
                if url:
                    new_uuid = datastore.add_watch(url=url,
                                                   extras=extras,
                                                   tag=tags,
                                                   write_to_disk_now=False)
                    if new_uuid:
                        # Straight into the queue.
                        self.new_uuids.append(new_uuid)
                        good += 1
        except Exception as e:
            logger.error(e)
            flash(f"Error processing row number {row_i}, check all cell data types are correct, row was skipped.", 'error')
        else:
            row_i += 1

        flash(
            "{} imported from custom .xlsx in {:.2f}s".format(len(self.new_uuids), time.time() - now))